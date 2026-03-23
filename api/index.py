import os
import csv
import io
import json
import secrets
from datetime import datetime, date
from functools import wraps
from typing import Any, cast

from flask import Flask, request, jsonify, session, send_file, Response
from sqlalchemy import create_engine, Column, Integer, String, Text, Date, DateTime, or_, func
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session
from werkzeug.security import generate_password_hash, check_password_hash

# ---------------------------------------------------------------------------
# App configuration
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///uict.db")
# Vercel Postgres may provide postgres:// which SQLAlchemy 2.x needs as postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = scoped_session(sessionmaker(bind=engine))
Base = declarative_base()

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class Admin(Base):
    __tablename__ = "admins"
    id = Column(Integer, primary_key=True, autoincrement=True)
    username = Column(String(80), unique=True, nullable=False)
    password_hash = Column(String(256), nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(cast(str, self.password_hash), password)


class Equipment(Base):
    __tablename__ = "equipment"
    id = Column(Integer, primary_key=True, autoincrement=True)
    equipment_type = Column(String(120), nullable=False)
    brand = Column(String(120), nullable=False)
    model = Column(String(120), nullable=False)
    serial_number = Column(String(120), unique=True, nullable=False)
    mr_to = Column(String(200), nullable=False)
    date_unserviceable = Column(Date, nullable=False)
    location = Column(String(200), nullable=False)
    remarks = Column(Text, default="")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    def to_dict(self):
        return {
            "id": self.id,
            "equipment_type": self.equipment_type,
            "brand": self.brand,
            "model": self.model,
            "serial_number": self.serial_number,
            "mr_to": self.mr_to,
            "date_unserviceable": self.date_unserviceable.isoformat() if self.date_unserviceable is not None else None,
            "location": self.location,
            "remarks": self.remarks or "",
            "created_at": self.created_at.isoformat() if self.created_at is not None else None,
            "updated_at": self.updated_at.isoformat() if self.updated_at is not None else None,
        }


# Create tables & seed default admin
Base.metadata.create_all(engine)

def _seed_admin():
    db = SessionLocal()
    try:
        if not db.query(Admin).first():
            admin = Admin(username="admin")
            admin.set_password("admin123")
            db.add(admin)
            db.commit()
    finally:
        db.close()

_seed_admin()

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_id"):
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated

# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    db = SessionLocal()
    try:
        admin = db.query(Admin).filter_by(username=username).first()
        if not admin or not admin.check_password(password):
            return jsonify({"error": "Invalid credentials"}), 401
        session["admin_id"] = admin.id
        session["admin_username"] = admin.username
        return jsonify({"message": "Login successful", "username": admin.username})
    finally:
        db.close()


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "Logged out"})


@app.route("/api/auth/check", methods=["GET"])
def auth_check():
    if session.get("admin_id"):
        return jsonify({"authenticated": True, "username": session.get("admin_username")})
    return jsonify({"authenticated": False}), 401

# ---------------------------------------------------------------------------
# Dashboard stats
# ---------------------------------------------------------------------------

@app.route("/api/dashboard/stats", methods=["GET"])
@login_required
def dashboard_stats():
    db = SessionLocal()
    try:
        total = db.query(func.count(Equipment.id)).scalar() or 0
        by_type = (
            db.query(Equipment.equipment_type, func.count(Equipment.id))
            .group_by(Equipment.equipment_type)
            .all()
        )
        by_location = (
            db.query(Equipment.location, func.count(Equipment.id))
            .group_by(Equipment.location)
            .all()
        )
        by_brand = (
            db.query(Equipment.brand, func.count(Equipment.id))
            .group_by(Equipment.brand)
            .all()
        )
        recent = (
            db.query(Equipment)
            .order_by(Equipment.created_at.desc())
            .limit(5)
            .all()
        )
        return jsonify({
            "total_equipment": total,
            "by_type": [{"label": r[0], "count": r[1]} for r in by_type],
            "by_location": [{"label": r[0], "count": r[1]} for r in by_location],
            "by_brand": [{"label": r[0], "count": r[1]} for r in by_brand],
            "recent": [e.to_dict() for e in recent],
        })
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Equipment CRUD
# ---------------------------------------------------------------------------

@app.route("/api/equipment", methods=["GET"])
@login_required
def list_equipment():
    db = SessionLocal()
    try:
        q = db.query(Equipment)

        # Filters
        search = request.args.get("search", "").strip()
        mr_to = request.args.get("mr_to", "").strip()
        location = request.args.get("location", "").strip()
        brand = request.args.get("brand", "").strip()
        equipment_type = request.args.get("equipment_type", "").strip()
        serial_number = request.args.get("serial_number", "").strip()

        if search:
            like = f"%{search}%"
            q = q.filter(
                or_(
                    Equipment.equipment_type.ilike(like),
                    Equipment.brand.ilike(like),
                    Equipment.model.ilike(like),
                    Equipment.serial_number.ilike(like),
                    Equipment.mr_to.ilike(like),
                    Equipment.location.ilike(like),
                    Equipment.remarks.ilike(like),
                )
            )
        if mr_to:
            q = q.filter(Equipment.mr_to.ilike(f"%{mr_to}%"))
        if location:
            q = q.filter(Equipment.location.ilike(f"%{location}%"))
        if brand:
            q = q.filter(Equipment.brand.ilike(f"%{brand}%"))
        if equipment_type:
            q = q.filter(Equipment.equipment_type.ilike(f"%{equipment_type}%"))
        if serial_number:
            q = q.filter(Equipment.serial_number.ilike(f"%{serial_number}%"))

        # Sorting
        sort_by = request.args.get("sort_by", "created_at")
        sort_dir = request.args.get("sort_dir", "desc")
        
        # Whitelist valid sort columns to prevent issues and improve security
        valid_sort_cols = {
            "created_at": Equipment.created_at,
            "equipment_type": Equipment.equipment_type,
            "brand": Equipment.brand,
            "model": Equipment.model,
            "serial_number": Equipment.serial_number,
            "mr_to": Equipment.mr_to,
            "date_unserviceable": Equipment.date_unserviceable,
            "location": Equipment.location,
            "updated_at": Equipment.updated_at
        }
        sort_col = cast(Any, valid_sort_cols.get(sort_by, Equipment.created_at))
        
        if sort_dir == "desc":
            q = q.order_by(sort_col.desc())
        else:
            q = q.order_by(sort_col.asc())

        # Pagination
        page = max(int(request.args.get("page", 1)), 1)
        per_page = min(int(request.args.get("per_page", 20)), 100)
        total = q.count()
        items = q.offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            "items": [e.to_dict() for e in items],
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max((total + per_page - 1) // per_page, 1),
        })
    finally:
        db.close()


@app.route("/api/equipment/<int:eid>", methods=["GET"])
@login_required
def get_equipment(eid):
    db = SessionLocal()
    try:
        eq = db.get(Equipment, eid)
        if not eq:
            return jsonify({"error": "Equipment not found"}), 404
        return jsonify(eq.to_dict())
    finally:
        db.close()


@app.route("/api/equipment", methods=["POST"])
@login_required
def create_equipment():
    data = request.get_json(silent=True) or {}
    required = ["equipment_type", "brand", "model", "serial_number", "mr_to", "date_unserviceable", "location"]
    missing = [f for f in required if not data.get(f, "").strip()]
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    db = SessionLocal()
    try:
        # Check duplicate serial
        if db.query(Equipment).filter_by(serial_number=data["serial_number"].strip()).first():
            return jsonify({"error": "Serial number already exists"}), 409

        eq = Equipment(
            equipment_type=data["equipment_type"].strip(),
            brand=data["brand"].strip(),
            model=data["model"].strip(),
            serial_number=data["serial_number"].strip(),
            mr_to=data["mr_to"].strip(),
            date_unserviceable=date.fromisoformat(data["date_unserviceable"]),
            location=data["location"].strip(),
            remarks=data.get("remarks", "").strip(),
        )
        db.add(eq)
        db.commit()
        db.refresh(eq)
        return jsonify(eq.to_dict()), 201
    except ValueError as e:
        return jsonify({"error": f"Invalid date format: {e}"}), 400
    finally:
        db.close()


@app.route("/api/equipment/<int:eid>", methods=["PUT"])
@login_required
def update_equipment(eid):
    data_raw = request.get_json(silent=True)
    if not isinstance(data_raw, dict):
        data_raw = {}
    data = cast(dict[str, Any], data_raw)
    
    db = SessionLocal()
    try:
        eq = db.get(Equipment, eid)
        if not eq:
            return jsonify({"error": "Equipment not found"}), 404

        # Check duplicate serial if changed
        new_serial = data.get("serial_number", "")
        new_serial = new_serial.strip() if isinstance(new_serial, str) else ""
        
        if new_serial and new_serial != eq.serial_number:
            # Important: exclude the current record from the check!
            if db.query(Equipment).filter(Equipment.serial_number == new_serial, Equipment.id != eid).first():
                return jsonify({"error": "Serial number already exists"}), 409

        for field in ["equipment_type", "brand", "model", "serial_number", "mr_to", "location", "remarks"]:
            if field in data:
                val = data.get(field)
                if isinstance(val, str):
                    setattr(eq, field, val.strip())

        if "date_unserviceable" in data:
            date_val = data.get("date_unserviceable")
            if isinstance(date_val, str):
                eq.date_unserviceable = cast(Any, date.fromisoformat(date_val))

        eq.updated_at = cast(Any, datetime.utcnow())
        db.commit()
        db.refresh(eq)
        return jsonify(eq.to_dict())
    except ValueError as e:
        return jsonify({"error": f"Invalid date format: {e}"}), 400
    finally:
        db.close()


@app.route("/api/equipment/<int:eid>", methods=["DELETE"])
@login_required
def delete_equipment(eid):
    db = SessionLocal()
    try:
        eq = db.get(Equipment, eid)
        if not eq:
            return jsonify({"error": "Equipment not found"}), 404
        db.delete(eq)
        db.commit()
        return jsonify({"message": "Equipment deleted"})
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Export — CSV
# ---------------------------------------------------------------------------

def _get_export_query(db):
    """Build filtered query for exports using the same filters as list."""
    q = db.query(Equipment)
    search = request.args.get("search", "").strip()
    mr_to = request.args.get("mr_to", "").strip()
    location = request.args.get("location", "").strip()
    brand = request.args.get("brand", "").strip()
    equipment_type = request.args.get("equipment_type", "").strip()
    serial_number = request.args.get("serial_number", "").strip()

    if search:
        like = f"%{search}%"
        q = q.filter(
            or_(
                Equipment.equipment_type.ilike(like),
                Equipment.brand.ilike(like),
                Equipment.model.ilike(like),
                Equipment.serial_number.ilike(like),
                Equipment.mr_to.ilike(like),
                Equipment.location.ilike(like),
            )
        )
    if mr_to:
        q = q.filter(Equipment.mr_to.ilike(f"%{mr_to}%"))
    if location:
        q = q.filter(Equipment.location.ilike(f"%{location}%"))
    if brand:
        q = q.filter(Equipment.brand.ilike(f"%{brand}%"))
    if equipment_type:
        q = q.filter(Equipment.equipment_type.ilike(f"%{equipment_type}%"))
    if serial_number:
        q = q.filter(Equipment.serial_number.ilike(f"%{serial_number}%"))
    return q.order_by(Equipment.created_at.desc())


@app.route("/api/equipment/export/csv", methods=["GET"])
@login_required
def export_csv():
    db = SessionLocal()
    try:
        items = _get_export_query(db).all()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "No.", "Equipment Type", "Brand", "Model", "Serial Number",
            "MR To", "Date Unserviceable", "Location", "Remarks"
        ])
        for i, eq in enumerate(items, 1):
            writer.writerow([
                i, eq.equipment_type, eq.brand, eq.model, eq.serial_number,
                eq.mr_to, eq.date_unserviceable.isoformat() if eq.date_unserviceable is not None else "",
                eq.location, eq.remarks or ""
            ])
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=uict_report_{date.today().isoformat()}.csv"}
        )
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Export — PDF
# ---------------------------------------------------------------------------

@app.route("/api/equipment/export/pdf", methods=["GET"])
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    db = SessionLocal()
    try:
        items = _get_export_query(db).all()
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16, spaceAfter=6)
        subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
        cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

        elements = []
        elements.append(Paragraph("Unserviceable ICT Equipment Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}  |  Total records: {len(items)}", subtitle_style))
        elements.append(Spacer(1, 12))

        header = ["No.", "Type", "Brand", "Model", "Serial No.", "MR To", "Date", "Location", "Remarks"]
        data: list[list[Any]] = [header]
        for i, eq in enumerate(items, 1):
            data.append([
                str(i),
                Paragraph(eq.equipment_type, cell_style),
                Paragraph(eq.brand, cell_style),
                Paragraph(eq.model, cell_style),
                Paragraph(eq.serial_number, cell_style),
                Paragraph(eq.mr_to, cell_style),
                eq.date_unserviceable.strftime("%Y-%m-%d") if eq.date_unserviceable is not None else "",
                Paragraph(eq.location, cell_style),
                Paragraph(eq.remarks or "", cell_style),
            ])

        col_widths = [0.35*inch, 0.9*inch, 0.8*inch, 0.9*inch, 1.1*inch, 1.2*inch, 0.8*inch, 1.2*inch, 1.5*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdfa")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"uict_report_{date.today().isoformat()}.pdf",
        )
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Distinct values for filter dropdowns
# ---------------------------------------------------------------------------

@app.route("/api/equipment/filters", methods=["GET"])
@login_required
def filter_options():
    db = SessionLocal()
    try:
        types = [r[0] for r in db.query(Equipment.equipment_type).distinct().order_by(Equipment.equipment_type).all()]
        brands = [r[0] for r in db.query(Equipment.brand).distinct().order_by(Equipment.brand).all()]
        locations = [r[0] for r in db.query(Equipment.location).distinct().order_by(Equipment.location).all()]
        people = [r[0] for r in db.query(Equipment.mr_to).distinct().order_by(Equipment.mr_to).all()]
        return jsonify({
            "equipment_types": types,
            "brands": brands,
            "locations": locations,
            "people": people,
        })
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Import — Excel (.xlsx)
# ---------------------------------------------------------------------------

@app.route("/api/equipment/import", methods=["POST"])
@login_required
def import_equipment():
    """Import equipment from an uploaded .xlsx file.

    Expected columns (case-insensitive, matched by header text):
        Equipment Type, Brand, Model, Serial Number,
        MR To, Date Declared Unserviceable (or Date Unserviceable), Location, Remarks

    Rows with duplicate serial numbers (already in DB) are skipped.
    """
    from openpyxl import load_workbook

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return jsonify({"error": "The workbook has no active sheet"}), 400

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"error": "The file has no data rows (only a header or empty)"}), 400

        # --- Map header columns ---
        raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        col_map = {}
        aliases = {
            "equipment_type": ["equipment type", "type", "equipment_type"],
            "brand": ["brand"],
            "model": ["model"],
            "serial_number": ["serial number", "serial no", "serial no.", "serial_number", "sn"],
            "mr_to": ["mr to", "mr_to", "person accountable", "accountable", "mr to (person accountable)"],
            "date_unserviceable": ["date declared unserviceable", "date unserviceable", "date_unserviceable", "date"],
            "location": ["location", "division", "field office", "location (division / field office)"],
            "remarks": ["remarks", "notes", "comment", "comments"],
        }
        for field, names in aliases.items():
            for idx, header in enumerate(raw_headers):
                if header in names:
                    col_map[field] = idx
                    break

        required_fields = ["equipment_type", "brand", "model", "serial_number", "mr_to", "date_unserviceable", "location"]
        missing = [f for f in required_fields if f not in col_map]
        if missing:
            friendly = {
                "equipment_type": "Equipment Type",
                "brand": "Brand",
                "model": "Model",
                "serial_number": "Serial Number",
                "mr_to": "MR To",
                "date_unserviceable": "Date Unserviceable",
                "location": "Location",
            }
            names = [friendly.get(m, m) for m in missing]
            return jsonify({"error": f"Missing required columns: {', '.join(names)}"}), 400

        db = SessionLocal()
        try:
            added = 0
            skipped = 0
            errors_list = []

            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    def cell(field):
                        idx = col_map.get(field)
                        if idx is None or idx >= len(row):
                            return ""
                        val = row[idx]
                        return str(val).strip() if val is not None else ""

                    serial = cell("serial_number")
                    if not serial:
                        errors_list.append(f"Row {row_num}: Empty serial number, skipped")
                        skipped += 1
                        continue

                    # Skip duplicates
                    if db.query(Equipment).filter_by(serial_number=serial).first():
                        skipped += 1
                        continue

                    date_str = cell("date_unserviceable")
                    parsed_date = None
                    if date_str:
                        # Try ISO format first
                        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%B %d, %Y"):
                            try:
                                parsed_date = datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        # openpyxl may return a datetime object directly
                        if parsed_date is None:
                            cell_val = row[col_map["date_unserviceable"]]
                            if hasattr(cell_val, "date"):
                                parsed_date = cell_val.date()
                            elif hasattr(cell_val, "year"):
                                parsed_date = date(cell_val.year, cell_val.month, cell_val.day)

                    if parsed_date is None:
                        errors_list.append(f"Row {row_num}: Invalid date '{date_str}', skipped")
                        skipped += 1
                        continue

                    eq = Equipment(
                        equipment_type=cell("equipment_type"),
                        brand=cell("brand"),
                        model=cell("model"),
                        serial_number=serial,
                        mr_to=cell("mr_to"),
                        date_unserviceable=parsed_date,
                        location=cell("location"),
                        remarks=cell("remarks") if "remarks" in col_map else "",
                    )
                    db.add(eq)
                    added += 1
                except Exception as row_err:
                    errors_list.append(f"Row {row_num}: {str(row_err)}")
                    skipped += 1

            db.commit()
            wb.close()

            result = {
                "message": f"Import complete: {added} added, {skipped} skipped",
                "added": added,
                "skipped": skipped,
            }
            if errors_list:
                result["errors"] = errors_list[:20]  # cap to avoid huge responses
            return jsonify(result), 200
        finally:
            db.close()

    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 400

# ---------------------------------------------------------------------------
# Change password
# ---------------------------------------------------------------------------

@app.route("/api/auth/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json(silent=True) or {}
    current = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    if not current or not new_pw:
        return jsonify({"error": "Both current and new password are required"}), 400
    if len(new_pw) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    db = SessionLocal()
    try:
        admin = db.get(Admin, session["admin_id"])
        if not admin or not admin.check_password(current):
            return jsonify({"error": "Current password is incorrect"}), 403
        admin.set_password(new_pw)
        db.commit()
        return jsonify({"message": "Password changed successfully"})
    finally:
        db.close()

# ---------------------------------------------------------------------------
# Serve frontend for local dev (not needed on Vercel)
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import pathlib

    static_dir = pathlib.Path(__file__).resolve().parent.parent / "public"

    @app.route("/")
    def serve_index():
        return send_file(static_dir / "index.html")

    @app.route("/css/<path:path>")
    def serve_css(path):
        return send_file(static_dir / "css" / path)

    @app.route("/js/<path:path>")
    def serve_js(path):
        return send_file(static_dir / "js" / path)

    app.run(debug=True, port=5000)
